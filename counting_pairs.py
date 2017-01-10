# -*- coding: utf-8 -*-
"""
Created on Sat Jul 02 23:59:51 2016

@author: Rahul
"""

import openpyxl

wb_J = openpyxl.load_workbook('jan.xlsx')
type(wb_J)
wb_J.get_sheet_names()
 
sheet_J_F = wb_J.get_sheet_by_name('Financial-Services')
sheet_J_F['A2'].value

jan = []
for i in range(2, sheet_J_F.max_row+1):
       a = [sheet_J_F.cell(row=i, column=1).value, sheet_J_F.cell(row=i, column=2).value]
       jan.append(a)
       

## loading the workbook
wb_J = openpyxl.load_workbook('jan.xlsx')
wb_F = openpyxl.load_workbook('feb.xlsx')
wb_M = openpyxl.load_workbook('march.xlsx')
wb_A = openpyxl.load_workbook('april.xlsx')
wb_MY = openpyxl.load_workbook('may.xlsx')
wb_Jn = openpyxl.load_workbook('june.xlsx')
wb_JY = openpyxl.load_workbook('july.xlsx')
wb_AT = openpyxl.load_workbook('august.xlsx')
wb_S = openpyxl.load_workbook('sep.xlsx')
wb_O = openpyxl.load_workbook('oct.xlsx')
wb_N = openpyxl.load_workbook('nov.xlsx')
wb_D = openpyxl.load_workbook('dec.xlsx')


name = "Industrial-Goods-&-Services"

# Checking for financial services
sheet_J_F = wb_J.get_sheet_by_name(name)
sheet_F_F = wb_F.get_sheet_by_name(name)
sheet_M_F = wb_M.get_sheet_by_name(name)
sheet_A_F = wb_A.get_sheet_by_name(name)
sheet_MY_F = wb_MY.get_sheet_by_name(name)
sheet_Jn_F = wb_Jn.get_sheet_by_name(name)
sheet_JY_F = wb_JY.get_sheet_by_name(name)
sheet_AT_F = wb_AT.get_sheet_by_name(name)
sheet_S_F = wb_S.get_sheet_by_name(name)
sheet_O_F = wb_O.get_sheet_by_name(name)
sheet_N_F = wb_N.get_sheet_by_name(name)
sheet_D_F = wb_D.get_sheet_by_name(name)

jan = []
for i in range(2, sheet_J_F.max_row+1):
       a = [sheet_J_F.cell(row=i, column=1).value, sheet_J_F.cell(row=i, column=2).value]
       jan.append(a)

feb = []
for i in range(2, sheet_F_F.max_row+1):
        b = [sheet_F_F.cell(row=i, column=1).value, sheet_F_F.cell(row=i, column=2).value]
        feb.append(b)
        
march = []        
for i in range(2, sheet_M_F.max_row+1):
        c = [sheet_M_F.cell(row=i, column=1).value, sheet_M_F.cell(row=i, column=2).value]
        march.append(c)
        
april = []        
for i in range(2, sheet_A_F.max_row+1):
        d = [sheet_A_F.cell(row=i, column=1).value, sheet_A_F.cell(row=i, column=2).value]
        april.append(d)

may = []
for i in range(2, sheet_MY_F.max_row+1):
        e = [sheet_MY_F.cell(row=i, column=1).value, sheet_MY_F.cell(row=i, column=2).value]
        may.append(e)


june = []
for i in range(2, sheet_Jn_F.max_row+1):
        f = [sheet_Jn_F.cell(row=i, column=1).value, sheet_Jn_F.cell(row=i, column=2).value]
        june.append(f)

july = []
for i in range(2, sheet_JY_F.max_row+1):
        g = [sheet_JY_F.cell(row=i, column=1).value, sheet_JY_F.cell(row=i, column=2).value]
        july.append(g)

august = []
for i in range(2, sheet_AT_F.max_row+1):
       h = [sheet_AT_F.cell(row=i, column=1).value, sheet_AT_F.cell(row=i, column=2).value]
       august.append(h)

sep = []
for i in range(2, sheet_S_F.max_row+1):
       i = [sheet_S_F.cell(row=i, column=1).value, sheet_S_F.cell(row=i, column=2).value]
       sep.append(i)

octr = []
for i in range(2, sheet_O_F.max_row+1):
       j = [sheet_O_F.cell(row=i, column=1).value, sheet_O_F.cell(row=i, column=2).value]
       octr.append(j)

nov = []       
for i in range(2, sheet_N_F.max_row+1):
       k = [sheet_N_F.cell(row=i, column=1).value, sheet_N_F.cell(row=i, column=2).value]
       nov.append(k)

dec = []
for i in range(2, sheet_D_F.max_row+1):
       l = [sheet_D_F.cell(row=i, column=1).value, sheet_D_F.cell(row=i, column=2).value] 
       dec.append(l)


L = jan+feb+march+april+may+june+july+august+sep+octr+nov+dec


####### counting particular sublist from list
s = [['HAVELLS', 'VOLTAS']]  ## result  = 6
s = [['ADANIPORTS', 'SIEMENS']] ## = 6
s = [['CROMPGREAV', 'VOLTAS']]  ##  = 4

sum(1 for i in range(len(L)) if L[i:i+len(s)]==s)

##########


## for counting unique element inside the list
from collections import Counter
Counter(L)

## counting unique list within list
Counter(str(e) for e in L)




















